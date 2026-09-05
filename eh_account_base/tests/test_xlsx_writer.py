# -*- encoding: utf-8 -*-
##############################################################################
#
# ERP Heritage
# Copyright (C) 2026 (https://www.erpheritage.com.au/)
#
##############################################################################
"""
Unit tests for XlsxReportWriter.

These tests do not require Odoo's DB, only openpyxl. They construct a
synthetic payload and verify that:

* The writer returns bytes that openpyxl can re open without errors.
* Title row, meta rows, header row, and data rows land in expected positions.
* Numeric cells get the correct number format and right alignment.
* Totals row carries bold font and a top border.
* Empty payloads do not crash.
* The sheet name is truncated to Excel's 31 character limit.
"""

from decimal import Decimal
import io

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from odoo.tests import tagged

from odoo.addons.eh_account_base.tools.xlsx_writer import XlsxReportWriter
from .common import EhAccountUnitTestCase


PAYLOAD = {
    'columns': [
        {'expression_label': 'account', 'name': 'Account',
         'figure_type': 'string'},
        {'expression_label': 'opening_debit', 'name': 'Opening DB',
         'figure_type': 'monetary'},
        {'expression_label': 'opening_credit', 'name': 'Opening CR',
         'figure_type': 'monetary'},
        {'expression_label': 'closing_debit', 'name': 'Closing DB',
         'figure_type': 'monetary'},
        {'expression_label': 'closing_credit', 'name': 'Closing CR',
         'figure_type': 'monetary'},
    ],
    'lines': [
        {
            'id': 'account-1',
            'name': '1000 Cash on Hand',
            'level': 1,
            'columns': [
                {'expression_label': 'opening_debit', 'value': 1000.0},
                {'expression_label': 'opening_credit', 'value': 0.0},
                {'expression_label': 'closing_debit', 'value': 1500.0},
                {'expression_label': 'closing_credit', 'value': 0.0},
            ],
        },
        {
            'id': 'account-2',
            'name': '4000 Sales Revenue',
            'level': 1,
            'columns': [
                {'expression_label': 'opening_debit', 'value': 0.0},
                {'expression_label': 'opening_credit', 'value': 1000.0},
                {'expression_label': 'closing_debit', 'value': 0.0},
                {'expression_label': 'closing_credit', 'value': 1500.0},
            ],
        },
    ],
    'totals': {
        'opening_debit': 1000.0,
        'opening_credit': 1000.0,
        'closing_debit': 1500.0,
        'closing_credit': 1500.0,
    },
    'meta': {
        'date_from': '2026-01-01',
        'date_to': '2026-12-31',
        'posted_only': True,
    },
    'generated_at': '2026-04-30T12:00:00',
}


@tagged('eh_account_base', 'unit')
class TestXlsxWriterShape(EhAccountUnitTestCase):

    def _render(self, payload=None, name="Trial Balance"):
        content = XlsxReportWriter(name).write_payload(payload or PAYLOAD)
        return content, load_workbook(io.BytesIO(content))

    def test_returns_xlsx_bytes(self):
        content, wb = self._render()
        self.assertIsInstance(content, bytes)
        # XLSX is a ZIP archive; magic bytes are 'PK'.
        self.assertEqual(content[:2], b'PK')
        self.assertEqual(wb.active.title, 'Report')

    def test_title_row_present(self):
        _, wb = self._render(name="My Custom Title")
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "My Custom Title")
        self.assertTrue(ws.cell(row=1, column=1).font.bold)

    def test_title_and_header_formula_injection_is_neutralised(self):
        payload = dict(PAYLOAD)
        payload['columns'] = [
            dict(PAYLOAD['columns'][0], name='=WEBSERVICE("bad")'),
            *PAYLOAD['columns'][1:],
        ]
        content = XlsxReportWriter(
            '+SUM(1,1)',
        ).write_payload(payload)
        ws = load_workbook(io.BytesIO(content)).active
        self.assertEqual(ws.cell(row=1, column=1).value, "'+SUM(1,1)")
        header_row = next(
            row for row in range(1, ws.max_row + 1)
            if ws.cell(row=row, column=1).value == "'=WEBSERVICE(\"bad\")"
        )
        self.assertGreater(header_row, 1)

    def test_meta_rows_show_period_and_filters(self):
        _, wb = self._render()
        ws = wb.active
        # Search the first 6 rows for the meta info.
        meta_text = " ".join(
            (ws.cell(row=r, column=1).value or '')
            for r in range(1, 7)
        )
        self.assertIn("2026-01-01", meta_text)
        self.assertIn("2026-12-31", meta_text)
        self.assertIn("Posted entries only", meta_text)
        meta_cell = next(
            ws.cell(row=row, column=1)
            for row in range(1, 7)
            if 'Posted entries only' in (
                ws.cell(row=row, column=1).value or ''
            )
        )
        self.assertTrue(meta_cell.alignment.wrap_text)
        self.assertEqual(meta_cell.alignment.vertical, 'top')
        self.assertGreaterEqual(
            ws.row_dimensions[meta_cell.row].height, 30,
        )

    def test_snapshot_meta_uses_as_at_without_blank_period_start(self):
        payload = dict(PAYLOAD)
        payload['meta'] = {
            'date_basis': 'as_of',
            'date_to': '2026-06-30',
            'posted_only': True,
        }
        _, wb = self._render(payload=payload, name="Balance Sheet")
        ws = wb.active
        meta_text = " ".join(
            (ws.cell(row=row, column=1).value or '')
            for row in range(1, 7)
        )
        self.assertIn("As at: 2026-06-30", meta_text)
        self.assertNotIn("Period:  to", meta_text)

    def test_column_headers_appear_with_correct_text(self):
        _, wb = self._render()
        ws = wb.active
        header_row = self._find_row_starting_with(ws, "Account")
        self.assertIsNotNone(header_row)
        self.assertEqual(ws.cell(row=header_row, column=1).value, "Account")
        self.assertEqual(ws.cell(row=header_row, column=2).value, "Opening DB")
        self.assertEqual(ws.cell(row=header_row, column=5).value, "Closing CR")

    def test_grouped_column_headers_merge_and_freeze_full_axis(self):
        payload = dict(PAYLOAD)
        payload['column_header_rows'] = [
            [
                {'name': 'Account', 'rowspan': 2, 'colspan': 1},
                {'name': 'Opening', 'rowspan': 1, 'colspan': 2},
                {'name': 'Closing', 'rowspan': 1, 'colspan': 2},
            ],
            [
                {'name': 'Debit'}, {'name': 'Credit'},
                {'name': 'Debit'}, {'name': 'Credit'},
            ],
        ]
        _, wb = self._render(payload=payload)
        ws = wb.active
        opening_row = next(
            row for row in range(1, ws.max_row + 1)
            if ws.cell(row=row, column=2).value == 'Opening'
        )
        self.assertIn(
            'A%s:A%s' % (opening_row, opening_row + 1),
            {str(cell_range) for cell_range in ws.merged_cells.ranges},
        )
        self.assertIn(
            'B%s:C%s' % (opening_row, opening_row),
            {str(cell_range) for cell_range in ws.merged_cells.ranges},
        )
        self.assertEqual(
            ws.cell(row=opening_row + 1, column=2).value, 'Debit',
        )
        self.assertTrue(
            ws.cell(row=opening_row + 1, column=2).alignment.wrap_text,
        )
        cash_row = self._find_row_starting_with(ws, '1000 Cash on Hand')
        freeze = ws.freeze_panes
        freeze_coordinate = (
            freeze.coordinate if hasattr(freeze, 'coordinate') else freeze
        )
        self.assertEqual(freeze_coordinate, 'B%s' % cash_row)

    def test_malformed_grouped_headers_fail_closed_to_flat_row(self):
        malformed_rows = (
            # Width overflow.
            [[{'name': 'Too wide', 'colspan': 6}]],
            # Hole under a row-spanning cell.
            [
                [{'name': 'Account', 'rowspan': 2},
                 {'name': 'Values', 'colspan': 4}],
                [{'name': 'Only one'}],
            ],
            # Boolean spans are not integers for this contract.
            [[{'name': 'Bad', 'colspan': True}]],
        )
        for header_rows in malformed_rows:
            with self.subTest(header_rows=header_rows):
                payload = dict(PAYLOAD)
                payload['column_header_rows'] = header_rows
                _, wb = self._render(payload=payload)
                ws = wb.active
                header_row = self._find_row_starting_with(ws, 'Account')
                self.assertEqual(
                    ws.cell(row=header_row, column=2).value, 'Opening DB',
                )
                self.assertFalse(ws.merged_cells.ranges)

    def test_grouped_header_formula_injection_is_neutralised(self):
        payload = dict(PAYLOAD)
        payload['column_header_rows'] = [[
            {'name': '=WEBSERVICE("bad")'},
            {'name': 'Values', 'colspan': 4},
        ]]
        _, wb = self._render(payload=payload)
        ws = wb.active
        self.assertIsNotNone(
            self._find_row_starting_with(ws, "'=WEBSERVICE(\"bad\")"),
        )
        self.assertFalse(any(
            cell.data_type == 'f'
            for row in ws.iter_rows()
            for cell in row
        ))

    def test_data_lines_carry_correct_values(self):
        _, wb = self._render()
        ws = wb.active
        cash_row = self._find_row_starting_with(ws, "1000 Cash on Hand")
        self.assertIsNotNone(cash_row)
        self.assertEqual(ws.cell(row=cash_row, column=2).value, 1000.0)
        self.assertEqual(ws.cell(row=cash_row, column=4).value, 1500.0)

    def test_monetary_cells_have_accounting_format(self):
        _, wb = self._render()
        ws = wb.active
        cash_row = self._find_row_starting_with(ws, "1000 Cash on Hand")
        money_cell = ws.cell(row=cash_row, column=2)
        self.assertIn('#,##0.00', money_cell.number_format)
        self.assertEqual(money_cell.alignment.horizontal, 'right')

    def test_totals_row_present_with_bold_font_and_top_border(self):
        _, wb = self._render()
        ws = wb.active
        totals_row = self._find_row_starting_with(ws, "Totals")
        self.assertIsNotNone(totals_row)
        totals_cell = ws.cell(row=totals_row, column=2)
        self.assertEqual(totals_cell.value, 1000.0)
        self.assertTrue(totals_cell.font.bold)
        self.assertIsNotNone(totals_cell.border.top.style)

    def test_partial_semantic_totals_do_not_emit_misleading_footer(self):
        payload = {
            'columns': [
                {'expression_label': 'account', 'name': 'Account',
                 'figure_type': 'string'},
                {'expression_label': 'amount', 'name': 'Current',
                 'figure_type': 'monetary'},
                {'expression_label': 'prior_amount', 'name': 'Prior',
                 'figure_type': 'monetary'},
            ],
            'lines': [{
                'id': 'net_profit', 'name': 'Net Profit',
                'columns': [
                    {'expression_label': 'amount', 'value': 12.0},
                    {'expression_label': 'prior_amount', 'value': 10.0},
                ],
            }],
            # Common report shape: only one display expression matches.
            'totals': {'amount': 12.0, 'prior_net_profit': 10.0},
        }
        content = XlsxReportWriter('Comparative').write_payload(payload)
        ws = load_workbook(io.BytesIO(content)).active
        self.assertIsNone(self._find_row_starting_with(ws, 'Totals'))

    def test_explicit_export_totals_drive_column_aligned_footer(self):
        payload = {
            'columns': [
                {'expression_label': 'account', 'name': 'Account',
                 'figure_type': 'string'},
                {'expression_label': 'amount', 'name': 'Current',
                 'figure_type': 'monetary'},
                {'expression_label': 'prior_amount', 'name': 'Prior',
                 'figure_type': 'monetary'},
            ],
            'lines': [],
            'totals': {'net_profit': 12.0, 'prior_net_profit': 10.0},
            'export_totals': {'amount': 12.0, 'prior_amount': 10.0},
        }
        content = XlsxReportWriter('Comparative').write_payload(payload)
        ws = load_workbook(io.BytesIO(content)).active
        totals_row = self._find_row_starting_with(ws, 'Totals')
        self.assertIsNotNone(totals_row)
        self.assertEqual(ws.cell(totals_row, 2).value, 12.0)
        self.assertEqual(ws.cell(totals_row, 3).value, 10.0)

    def test_malformed_numeric_totals_are_suppressed(self):
        columns = [
            {'expression_label': 'account', 'name': 'Account',
             'figure_type': 'string'},
            {'expression_label': 'amount', 'name': 'Current',
             'figure_type': 'monetary'},
            {'expression_label': 'prior_amount', 'name': 'Prior',
             'figure_type': 'monetary'},
        ]
        malformed = [
            {'amount': 12.0},
            {'amount': None, 'prior_amount': 10.0},
            {'amount': True, 'prior_amount': 10.0},
            {'amount': float('nan'), 'prior_amount': 10.0},
            {'amount': 10 ** 10000, 'prior_amount': 10.0},
            {'amount': Decimal('1e10000'), 'prior_amount': 10.0},
        ]
        for totals_key in ('totals', 'export_totals'):
            for total_map in malformed:
                with self.subTest(
                    totals_key=totals_key, total_map=total_map,
                ):
                    payload = {
                        'columns': columns,
                        'lines': [],
                        totals_key: total_map,
                    }
                    content = XlsxReportWriter(
                        'Malformed totals',
                    ).write_payload(payload)
                    ws = load_workbook(io.BytesIO(content)).active
                    self.assertIsNone(
                        self._find_row_starting_with(ws, 'Totals'),
                    )

    def test_formula_like_numeric_totals_never_enter_workbook(self):
        columns = [
            {'expression_label': 'account', 'name': 'Account',
             'figure_type': 'string'},
            {'expression_label': 'amount', 'name': 'Current',
             'figure_type': 'monetary'},
            {'expression_label': 'prior_amount', 'name': 'Prior',
             'figure_type': 'monetary'},
        ]
        for totals_key in ('totals', 'export_totals'):
            with self.subTest(totals_key=totals_key):
                payload = {
                    'columns': columns,
                    'lines': [],
                    totals_key: {
                        'amount': '=WEBSERVICE("https://invalid.example")',
                        'prior_amount': 10.0,
                    },
                }
                content = XlsxReportWriter(
                    'Formula-safe totals',
                ).write_payload(payload)
                ws = load_workbook(io.BytesIO(content), data_only=False).active
                self.assertIsNone(self._find_row_starting_with(ws, 'Totals'))
                self.assertFalse(any(
                    cell.data_type == 'f'
                    for row in ws.iter_rows()
                    for cell in row
                ))

    def test_numeric_total_columns_require_unique_labels(self):
        payload = {
            'columns': [
                {'expression_label': 'account', 'name': 'Account',
                 'figure_type': 'string'},
                {'expression_label': 'amount', 'name': 'Current',
                 'figure_type': 'monetary'},
                {'expression_label': 'amount', 'name': 'Prior',
                 'figure_type': 'monetary'},
            ],
            'lines': [],
            'export_totals': {'amount': 12.0},
        }
        content = XlsxReportWriter('Duplicate labels').write_payload(payload)
        ws = load_workbook(io.BytesIO(content)).active
        self.assertIsNone(self._find_row_starting_with(ws, 'Totals'))

    def test_column_widths_set(self):
        _, wb = self._render()
        ws = wb.active
        # First column (name) is wider than monetary columns.
        first_width = ws.column_dimensions[get_column_letter(1)].width
        money_width = ws.column_dimensions[get_column_letter(2)].width
        self.assertGreater(first_width, money_width)

    def test_monetary_columns_expand_for_full_period_headers(self):
        current = '2025-07-01 to 2026-06-30'
        prior = '2024-07-01 to 2025-06-30'
        content = XlsxReportWriter('Profit and Loss').write_payload({
            'columns': [
                {'expression_label': 'account', 'name': 'Account',
                 'figure_type': 'string'},
                {'expression_label': 'amount', 'name': current,
                 'figure_type': 'monetary'},
                {'expression_label': 'prior_amount', 'name': prior,
                 'figure_type': 'monetary'},
            ],
            'lines': [],
            'totals': {},
            'generated_at': '',
        })
        ws = load_workbook(io.BytesIO(content)).active
        self.assertGreaterEqual(ws.column_dimensions['B'].width,
                                len(current) + 2)
        self.assertGreaterEqual(ws.column_dimensions['C'].width,
                                len(prior) + 2)
        self.assertTrue(ws.sheet_properties.pageSetUpPr.fitToPage)
        self.assertEqual(ws.page_setup.fitToWidth, 1)

    def test_grouped_columns_size_from_leaf_and_repeat_header_band(self):
        payload = self._grouped_axis_payload(6)
        _, wb = self._render(payload=payload, name='Grouped comparison')
        ws = wb.active
        header_row = self._find_row_starting_with(ws, 'Account')

        # The flattened labels deliberately contain the complete period and
        # analytic path. Only the short visible leaf label should size each
        # grouped value column.
        for column_index in range(2, 8):
            self.assertEqual(
                ws.column_dimensions[get_column_letter(column_index)].width,
                16,
            )
        self.assertEqual(ws.page_setup.orientation, 'landscape')
        self.assertEqual(str(ws.page_setup.paperSize), '9')
        self.assertFalse(ws.sheet_properties.pageSetUpPr.fitToPage)
        self.assertFalse(ws.sheet_properties.pageSetUpPr.autoPageBreaks)
        self.assertIsNone(ws.page_setup.fitToHeight)
        self.assertIsNone(ws.page_setup.fitToWidth)
        self.assertEqual(ws.page_setup.scale, 90)
        self.assertFalse(ws.col_breaks.brk)
        self.assertEqual(
            ws.print_title_rows.replace('$', ''),
            '%s:%s' % (header_row, header_row + 2),
        )
        self.assertEqual(ws.print_title_cols.replace('$', ''), 'A:A')

    def test_wide_grouped_columns_use_legible_horizontal_page_bands(self):
        payload = self._grouped_axis_payload(36)
        _, wb = self._render(payload=payload, name='Wide grouped comparison')
        ws = wb.active

        # Six normally sized value columns are allocated per landscape page,
        # with explicit breaks aligned to the analytic groups. Parent period
        # merges are split and their label repeated, so no merged text crosses
        # a physical print band in Excel or LibreOffice.
        self.assertIsNone(ws.page_setup.fitToWidth)
        self.assertEqual(
            [page_break.id for page_break in ws.col_breaks.brk],
            [7, 13, 19, 25, 31],
        )
        for column_index in range(2, 38):
            self.assertEqual(
                ws.column_dimensions[get_column_letter(column_index)].width,
                16,
            )
        header_row = self._find_row_starting_with(ws, 'Account')
        merged_ranges = {
            str(cell_range) for cell_range in ws.merged_cells.ranges
        }
        for first_column, last_column in (
            ('B', 'G'), ('H', 'M'), ('N', 'S'),
            ('T', 'Y'), ('Z', 'AE'), ('AF', 'AK'),
        ):
            self.assertIn(
                '%s%s:%s%s' % (
                    first_column, header_row, last_column, header_row,
                ),
                merged_ranges,
            )
        self.assertEqual(
            [
                ws.cell(header_row, column_index).value
                for column_index in (2, 8, 14, 20, 26, 32)
            ],
            [
                '2024-07-01 to 2025-06-30',
                '2024-07-01 to 2025-06-30',
                '2024-07-01 to 2025-06-30',
                '2025-07-01 to 2026-06-30',
                '2025-07-01 to 2026-06-30',
                '2025-07-01 to 2026-06-30',
            ],
        )
        self.assertEqual(
            [
                ws.cell(header_row + 1, column_index).value
                for column_index in (2, 8, 14, 20, 26, 32)
            ],
            [
                'Analytic 1', 'Analytic 2', 'Analytic 3',
                'Analytic 1', 'Analytic 2', 'Analytic 3',
            ],
        )

    def test_empty_payload_does_not_crash(self):
        content = XlsxReportWriter("Empty").write_payload({
            'columns': [],
            'lines': [],
            'totals': {},
            'generated_at': '',
        })
        self.assertIsInstance(content, bytes)

    def test_sheet_name_truncated_to_31_chars(self):
        very_long = "A" * 50
        writer = XlsxReportWriter("Report", sheet_name=very_long)
        # openpyxl raises if title exceeds 31 chars; survival is the assertion.
        content = writer.write_payload(PAYLOAD)
        wb = load_workbook(io.BytesIO(content))
        self.assertLessEqual(len(wb.active.title), 31)

    def test_negative_values_render(self):
        payload = {
            'columns': [
                {'expression_label': 'account', 'name': 'Account',
                 'figure_type': 'string'},
                {'expression_label': 'balance', 'name': 'Balance',
                 'figure_type': 'monetary'},
            ],
            'lines': [
                {'id': 'l1', 'name': 'Loss', 'level': 1,
                 'columns': [{'expression_label': 'balance', 'value': -1234.56}]},
            ],
            'totals': {'balance': -1234.56},
            'generated_at': '',
        }
        content, wb = self._render(payload=payload)
        ws = wb.active
        loss_row = self._find_row_starting_with(ws, "Loss")
        self.assertEqual(ws.cell(row=loss_row, column=2).value, -1234.56)

    def test_line_cell_figure_type_overrides_mixed_value_column(self):
        payload = {
            'columns': [
                {'expression_label': 'metric', 'name': 'Metric',
                 'figure_type': 'string'},
                {'expression_label': 'value', 'name': 'Value',
                 'figure_type': 'string'},
            ],
            'lines': [
                {'id': 'money', 'name': 'Revenue', 'level': 1,
                 'columns': [{
                     'expression_label': 'value', 'value': 1234.56,
                     'figure_type': 'monetary',
                 }]},
                {'id': 'ratio', 'name': 'Margin', 'level': 1,
                 'columns': [{
                     'expression_label': 'value', 'value': 0.25,
                     'figure_type': 'percentage',
                 }]},
                {'id': 'undefined', 'name': 'Undefined', 'level': 1,
                 'columns': [{
                     'expression_label': 'value', 'value': 'n/a',
                     'figure_type': 'percentage',
                 }]},
            ],
            'totals': {},
            'currency': {
                'name': 'AUD', 'symbol': '$', 'position': 'before',
                'decimal_places': 2,
            },
            'generated_at': '',
        }
        _, wb = self._render(payload=payload)
        ws = wb.active
        money = ws.cell(
            row=self._find_row_starting_with(ws, 'Revenue'), column=2,
        )
        ratio = ws.cell(
            row=self._find_row_starting_with(ws, 'Margin'), column=2,
        )
        undefined = ws.cell(
            row=self._find_row_starting_with(ws, 'Undefined'), column=2,
        )
        self.assertEqual(money.value, 1234.56)
        self.assertIn('$', money.number_format)
        self.assertEqual(money.alignment.horizontal, 'right')
        self.assertEqual(ratio.value, 0.25)
        self.assertEqual(ratio.number_format, '0.00%')
        self.assertEqual(undefined.value, 'n/a')

    def test_zero_decimal_currency_format_has_no_fraction(self):
        writer = XlsxReportWriter("Zero-decimal report")
        writer._currency = {
            'symbol': '¥',
            'position': 'before',
            'decimal_places': 0,
            'multi_currency': False,
        }
        number_format = writer._monetary_format()
        self.assertIn('#,##0', number_format)
        self.assertNotIn('.00', number_format)

    @staticmethod
    def _grouped_axis_payload(value_column_count):
        columns = [{
            'expression_label': 'account',
            'name': 'Account',
            'figure_type': 'string',
        }]
        for column_index in range(1, value_column_count + 1):
            period_index = (column_index - 1) // 18
            analytic_index = ((column_index - 1) // 6) % 3 + 1
            period = (
                '2024-07-01 to 2025-06-30'
                if period_index == 0
                else '2025-07-01 to 2026-06-30'
            )
            columns.append({
                'expression_label': 'amount_%s' % column_index,
                'name': (
                    '%s / Operating Plan / Analytic Account %02d / Value'
                ) % (period, analytic_index),
                'figure_type': 'monetary',
            })
        period_groups = []
        for period_start in range(0, value_column_count, 18):
            period_groups.append({
                'name': (
                    '2024-07-01 to 2025-06-30'
                    if period_start == 0
                    else '2025-07-01 to 2026-06-30'
                ),
                'colspan': min(18, value_column_count - period_start),
            })
        analytic_groups = []
        for analytic_start in range(0, value_column_count, 6):
            analytic_groups.append({
                'name': 'Analytic %s' % (
                    (analytic_start // 6) % 3 + 1
                ),
                'colspan': min(6, value_column_count - analytic_start),
            })
        return {
            'columns': columns,
            'column_header_rows': [
                [
                    {'name': 'Account', 'rowspan': 3},
                    *period_groups,
                ],
                analytic_groups,
                [
                    {'name': 'Value'}
                    for _unused in range(value_column_count)
                ],
            ],
            'lines': [],
            'totals': {},
            'generated_at': '',
        }

    @staticmethod
    def _find_row_starting_with(ws, text):
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == text:
                return r
        return None
