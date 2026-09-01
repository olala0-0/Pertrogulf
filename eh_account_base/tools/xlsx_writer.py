# -*- encoding: utf-8 -*-
##############################################################################
#
# ERP Heritage
# Copyright (C) 2026 (https://www.erpheritage.com.au/)
#
##############################################################################
"""
XLSX writer for dynamic report payloads.

Consumes the standard payload shape produced by every dynamic report
handler:

    {
        'columns': [
            {'expression_label': str, 'name': str, 'figure_type': str},
            ...
        ],
        'column_header_rows': [
            [
                {'name': str, 'colspan': int, 'rowspan': int},
                ...
            ],
            ...
        ],  # optional; exact rectangular cover of flat columns
        'lines': [
            {
                'id': str, 'name': str, 'level': int,
                'columns': [{'expression_label': str, 'value': any}, ...],
                'unfoldable': bool,
                'meta': dict,
            },
            ...
        ],
        'totals': {expression_label: number, ...},
        'meta': {date_from, date_to, posted_only, ...},
        'generated_at': iso datetime,
    }

Convention: payload['columns'][0] is always the line label column, rendered
from line['name']. payload['columns'][1:] are the value columns, rendered
from line['columns'][i] where i is the value column index.

Output is a single sheet workbook with:

* A title row with the report name.
* Optional metadata rows (period, posted only flag, generated at).
* One flat header row, or validated grouped/merged header rows.
* One data row per line.
* A totals row with a top border, when totals are present.
* Frozen header rows and first label column for wide matrices.

The writer is intentionally a plain Python class so unit tests do not need
the full Odoo registry. openpyxl is the only external dependency, and it is
already a standard Odoo requirement.
"""

from decimal import Decimal
import io
import math

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


# Number formats for openpyxl. The accounting convention is to render
# negatives in red and positive numbers without a sign.
_FORMAT_MONETARY = '#,##0.00;[Red](#,##0.00)'
_FORMAT_INTEGER = '#,##0'
_FORMAT_FLOAT = '#,##0.0000'
_FORMAT_PERCENT = '0.00%'
_FORMAT_DATE = 'yyyy-mm-dd'
_FORMAT_DATETIME = 'yyyy-mm-dd hh:mm:ss'

_NUMERIC_FIGURE_TYPES = frozenset({'monetary', 'integer', 'float', 'percentage'})

# A landscape A4 page can carry roughly six normally sized monetary columns
# beside the repeated account/line label without shrinking the text into an
# unreadable matrix. Wider reports deliberately use additional horizontal
# pages; every page repeats the label column and complete heading band.
_PRINT_VALUE_WIDTH_PER_PAGE = 6 * 16


class XlsxReportWriter:
    """Render a dynamic report payload to XLSX bytes.

    Use:

        writer = XlsxReportWriter("Trial Balance")
        content_bytes = writer.write_payload(payload)

    The writer is single use. Construct a new instance per render.
    """

    DEFAULT_FONT = Font(name='Arial', size=10)
    HEADER_FONT = Font(name='Arial', size=10, bold=True)
    TITLE_FONT = Font(name='Arial', size=14, bold=True)
    META_FONT = Font(name='Arial', size=9, italic=True, color='666666')
    TOTAL_FONT = Font(name='Arial', size=10, bold=True)
    HEADER_FILL = PatternFill(
        start_color='F0F0F0', end_color='F0F0F0', fill_type='solid',
    )
    TOP_BORDER = Border(top=Side(border_style='thin', color='000000'))

    def __init__(self, report_name, sheet_name='Report'):
        self.report_name = report_name or "Report"
        # Excel imposes a 31 character maximum on sheet names.
        self.sheet_name = (sheet_name or "Report")[:31]
        self.workbook = Workbook()
        self.ws = self.workbook.active
        self.ws.title = self.sheet_name
        self._row = 1
        self._currency = None

    BRAND_FONT = Font(
        name='Arial', size=8, italic=True, color='6C757D',
    )
    BRAND_FOOTER_TEXT = "Made with 🤍 from Melbourne by ERP Heritage"

    def write_payload(self, payload, meta=None):
        meta = meta if meta is not None else (payload.get('meta') or {})
        generated_at = payload.get('generated_at') or ''
        self._currency = payload.get('currency') or None

        self._write_title()
        self._write_meta(meta, generated_at)
        self._write_blank_row()

        columns = payload.get('columns') or []
        header_rows = self._normalise_column_header_rows(
            columns, payload.get('column_header_rows'),
        )
        self._set_column_widths(columns, header_rows)
        print_bands = self._grouped_print_bands(columns, header_rows)
        rendered_header_rows = self._split_header_rows_for_print(
            header_rows, print_bands,
        )
        header_first_row = self._row
        self._write_column_headers(columns, rendered_header_rows)
        header_last_row = self._row - 1
        # Keep every heading row and the account/line label visible while a
        # wide period x analytic matrix scrolls. ``self._row`` is now the
        # first data row for both flat and grouped headers.
        if columns:
            self.ws.freeze_panes = self.ws.cell(row=self._row, column=2)
        for line in (payload.get('lines') or []):
            self._write_line(line, columns)
        export_totals = self._resolve_export_totals(payload, columns)
        if export_totals is not None:
            self._write_totals(export_totals, columns)

        self._configure_print_layout(
            columns, header_first_row, header_last_row,
            rendered_header_rows, print_bands,
        )
        self._write_brand_footer()
        return self._serialise()

    def _write_brand_footer(self):
        """Write the ERP Heritage Melbourne brand line as the last
        non-data row of the sheet. Single cell, italic, muted grey, so
        it does not compete with the report content above.
        """
        self._row += 1
        cell = self.ws.cell(
            row=self._row, column=1, value=self.BRAND_FOOTER_TEXT,
        )
        cell.font = self.BRAND_FONT
        cell.alignment = Alignment(horizontal='left')

    # ---- title and meta rows ----

    def _write_title(self):
        ws = self.ws
        cell = ws.cell(
            row=self._row,
            column=1,
            value=self._safe_cell_text(self.report_name),
        )
        cell.font = self.TITLE_FONT
        self._row += 1

    def _write_meta(self, meta, generated_at):
        ws = self.ws
        details = []
        date_from = meta.get('date_from')
        date_to = meta.get('date_to')
        if meta.get('date_basis') == 'as_of' and date_to:
            details.append("As at: %s" % date_to)
        elif date_from or date_to:
            details.append("Period: %s to %s" % (date_from or '', date_to or ''))
        if 'posted_only' in meta:
            details.append(
                "Posted entries only" if meta['posted_only']
                else "All entries (including draft)",
            )
        if 'show_zero' in meta and meta['show_zero']:
            details.append("Including zero balance accounts")
        if self._currency:
            if self._currency.get('multi_currency'):
                details.append("Multi-currency scope")
            elif self._currency.get('name'):
                sym = self._currency.get('symbol') or ''
                details.append(
                    "Currency: %s%s" % (
                        self._currency['name'],
                        " (%s)" % sym if sym else "",
                    ),
                )
        if details:
            detail_text = " | ".join(details)
            cell = ws.cell(
                row=self._row, column=1, value=detail_text,
            )
            cell.font = self.META_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            # The label column repeats on every horizontal print band. Text
            # that merely overflows into adjacent blank cells on band one is
            # clipped on later bands, so give the wrapped metadata a stable
            # height based on the visible 35-character label column.
            wrapped_lines = max(1, int(math.ceil(len(detail_text) / 45.0)))
            ws.row_dimensions[self._row].height = wrapped_lines * 15
            self._row += 1
        if generated_at:
            ws.cell(
                row=self._row, column=1,
                value="Generated: %s" % generated_at,
            ).font = self.META_FONT
            self._row += 1

    def _write_blank_row(self):
        self._row += 1

    # ---- column header row ----

    def _write_column_headers(self, columns, header_rows=None):
        if header_rows:
            self._write_grouped_column_headers(columns, header_rows)
            return
        self._write_flat_column_headers(columns)

    def _write_flat_column_headers(self, columns):
        ws = self.ws
        for i, col_def in enumerate(columns, start=1):
            cell = ws.cell(
                row=self._row,
                column=i,
                value=self._safe_cell_text(col_def.get('name', '')),
            )
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(
                horizontal=(
                    'right'
                    if col_def.get('figure_type') in _NUMERIC_FIGURE_TYPES
                    else 'left'
                ),
                vertical='center',
                wrap_text=True,
            )
        self._row += 1

    @staticmethod
    def _normalise_column_header_rows(columns, raw_rows):
        """Return a rectangular header grid or ``None``.

        Flat ``columns`` remain authoritative. Optional grouped metadata is
        accepted only when positive integer spans cover that exact width with
        no overlap or hole. A stale/custom payload therefore falls back to
        the proven single-row export instead of producing shifted headings.
        """
        if (
            not isinstance(columns, list) or not columns
            or not isinstance(raw_rows, list) or not raw_rows
        ):
            return None
        height = len(raw_rows)
        width = len(columns)
        occupied = [[False] * width for _unused in range(height)]
        normalised = []
        for row_index, raw_row in enumerate(raw_rows):
            if not isinstance(raw_row, list) or not raw_row:
                return None
            row = []
            cursor = 0
            for raw_cell in raw_row:
                while cursor < width and occupied[row_index][cursor]:
                    cursor += 1
                if not isinstance(raw_cell, dict):
                    return None
                colspan = raw_cell.get('colspan', 1)
                rowspan = raw_cell.get('rowspan', 1)
                if (
                    isinstance(colspan, bool) or isinstance(rowspan, bool)
                    or not isinstance(colspan, int)
                    or not isinstance(rowspan, int)
                    or colspan < 1 or rowspan < 1
                    or cursor + colspan > width
                    or row_index + rowspan > height
                ):
                    return None
                for y_pos in range(row_index, row_index + rowspan):
                    for x_pos in range(cursor, cursor + colspan):
                        if occupied[y_pos][x_pos]:
                            return None
                        occupied[y_pos][x_pos] = True
                name = raw_cell.get('name', '')
                if (
                    isinstance(name, bool)
                    or not isinstance(name, (str, int, float))
                    or (
                        isinstance(name, float)
                        and not math.isfinite(name)
                    )
                ):
                    return None
                row.append({
                    'name': str(name),
                    'colspan': colspan,
                    'rowspan': rowspan,
                    'start': cursor,
                })
                cursor += colspan
            normalised.append(row)
        if any(not value for row in occupied for value in row):
            return None
        return normalised

    def _write_grouped_column_headers(self, columns, header_rows):
        ws = self.ws
        first_row = self._row
        merge_ranges = []
        for row_offset, header_row in enumerate(header_rows):
            sheet_row = first_row + row_offset
            for header_cell in header_row:
                first_col = header_cell['start'] + 1
                last_col = first_col + header_cell['colspan'] - 1
                last_row = sheet_row + header_cell['rowspan'] - 1
                # Style the complete physical range before merging. This
                # preserves fill/borders in Excel and LibreOffice.
                for row_index in range(sheet_row, last_row + 1):
                    for col_index in range(first_col, last_col + 1):
                        cell = ws.cell(row=row_index, column=col_index)
                        cell.font = self.HEADER_FONT
                        cell.fill = self.HEADER_FILL
                        cell.alignment = Alignment(
                            horizontal=(
                                'left' if first_col == 1 else 'center'
                            ),
                            vertical='center',
                            wrap_text=True,
                        )
                ws.cell(
                    row=sheet_row, column=first_col,
                    value=self._safe_cell_text(header_cell['name']),
                )
                if last_col != first_col or last_row != sheet_row:
                    merge_ranges.append(
                        (sheet_row, first_col, last_row, last_col),
                    )
        for first_row_index, first_col, last_row, last_col in merge_ranges:
            ws.merge_cells(
                start_row=first_row_index, start_column=first_col,
                end_row=last_row, end_column=last_col,
            )
        self._row += len(header_rows)

    # ---- data lines ----

    @staticmethod
    def _safe_cell_text(value):
        """Neutralise spreadsheet formula injection.

        Account/partner/ref text is free-form and user-controlled. Excel and
        LibreOffice auto-execute a cell whose text begins with '=', '+', '-'
        or '@' as a formula (e.g. a partner named =HYPERLINK(...)). Prefix
        such string values with an apostrophe so they render as literal text.
        Numbers are returned unchanged.
        """
        if isinstance(value, str) and value[:1] in ('=', '+', '-', '@'):
            return "'" + value
        return value

    def _write_line(self, line, columns):
        ws = self.ws
        # Column 1: line name. Indent if this is a sub line.
        name_cell = ws.cell(
            row=self._row, column=1,
            value=self._safe_cell_text(line.get('name', '')),
        )
        name_cell.font = (
            self.TOTAL_FONT if line.get('level', 1) == 0 else self.DEFAULT_FONT
        )
        level = max(int(line.get('level', 1)) - 1, 0)
        if level:
            name_cell.alignment = Alignment(indent=level)

        # Subsequent columns: line.columns[i] -> sheet column i + 2.
        for i, line_col in enumerate(line.get('columns') or []):
            sheet_col = i + 2
            col_def = columns[i + 1] if i + 1 < len(columns) else {}
            value = self._safe_cell_text(line_col.get('value'))
            cell = ws.cell(row=self._row, column=sheet_col, value=value)
            cell.font = self.DEFAULT_FONT
            figure_type = (
                line_col.get('figure_type')
                or col_def.get('figure_type', 'string')
            )
            self._apply_figure_type(cell, figure_type)

        self._row += 1

    # ---- totals row ----

    @staticmethod
    def _resolve_export_totals(payload, columns):
        """Return a complete column-aligned footer, or suppress it.

        Handler ``totals`` dictionaries often use semantic keys such as
        ``net_profit`` or ``total_equity`` rather than visible column
        expressions such as ``amount`` / ``prior_amount``. Writing a partly
        matched generic footer produces authoritative-looking blanks. An
        explicit ``export_totals`` map is preferred; otherwise legacy totals
        are used only when every numeric display column has a unique matching
        key containing a finite numeric scalar.  Returning a sanitised map
        also prevents a formula-like string from reaching openpyxl as an
        executable totals cell.  Section/computed total lines already remain
        in the workbook when this redundant footer is suppressed.
        """
        if 'export_totals' in payload:
            totals = payload.get('export_totals')
        else:
            totals = payload.get('totals')
        if not isinstance(totals, dict) or not totals:
            return None

        numeric_labels = []
        for column in columns[1:]:
            if column.get('figure_type') not in _NUMERIC_FIGURE_TYPES:
                continue
            label = column.get('expression_label')
            if (
                not isinstance(label, str)
                or not label
                or label in numeric_labels
            ):
                return None
            numeric_labels.append(label)
        if not numeric_labels:
            return None

        resolved = {}
        for label in numeric_labels:
            value = totals.get(label)
            if not XlsxReportWriter._is_supported_total_value(value):
                return None
            resolved[label] = value
        return resolved

    @staticmethod
    def _is_supported_total_value(value):
        """Whether openpyxl can safely write a numeric footer scalar."""
        if isinstance(value, bool) or not isinstance(
            value, (int, float, Decimal),
        ):
            return False
        try:
            # XLSX numeric cells are IEEE-754 doubles.  ``Decimal`` and
            # arbitrary-size ``int`` values can be mathematically finite yet
            # still overflow that storage type, so validate the same
            # conversion openpyxl/Excel ultimately has to perform.
            return math.isfinite(float(value))
        except (TypeError, ValueError, OverflowError):
            return False

    def _write_totals(self, totals, columns):
        ws = self.ws
        label_cell = ws.cell(row=self._row, column=1, value="Totals")
        label_cell.font = self.TOTAL_FONT
        label_cell.border = self.TOP_BORDER

        for i, col_def in enumerate(columns[1:], start=2):
            label = col_def.get('expression_label')
            value = totals.get(label) if label else None
            cell = ws.cell(row=self._row, column=i, value=value)
            cell.font = self.TOTAL_FONT
            cell.border = self.TOP_BORDER
            self._apply_figure_type(cell, col_def.get('figure_type', 'string'))

        self._row += 1

    # ---- formatting and widths ----

    def _monetary_format(self):
        """Build the Excel number format string from the payload currency.

        When the payload has no currency block (legacy or multi_currency),
        fall back to the static format that just renders thousands and
        two decimals.
        """
        if not self._currency or self._currency.get('multi_currency'):
            return _FORMAT_MONETARY
        raw_decimals = self._currency.get('decimal_places')
        decimals = int(2 if raw_decimals is None else raw_decimals)
        symbol = self._currency.get('symbol') or ''
        symbol_lit = symbol.replace('"', '\\"')
        if decimals <= 0:
            digits = '0'
        else:
            digits = '0.' + ('0' * decimals)
        if not symbol_lit:
            return f'#,##{digits};[Red](#,##{digits})'
        if self._currency.get('position') == 'before':
            return (
                f'"{symbol_lit}" #,##{digits};'
                f'[Red]"{symbol_lit}" (#,##{digits})'
            )
        return (
            f'#,##{digits} "{symbol_lit}";'
            f'[Red](#,##{digits}) "{symbol_lit}"'
        )

    def _apply_figure_type(self, cell, figure_type):
        if figure_type == 'monetary':
            cell.number_format = self._monetary_format()
            cell.alignment = Alignment(horizontal='right')
        elif figure_type == 'integer':
            cell.number_format = _FORMAT_INTEGER
            cell.alignment = Alignment(horizontal='right')
        elif figure_type == 'float':
            cell.number_format = _FORMAT_FLOAT
            cell.alignment = Alignment(horizontal='right')
        elif figure_type == 'percentage':
            cell.number_format = _FORMAT_PERCENT
            cell.alignment = Alignment(horizontal='right')
        elif figure_type == 'date':
            cell.number_format = _FORMAT_DATE
        elif figure_type == 'datetime':
            cell.number_format = _FORMAT_DATETIME
        # 'string', 'boolean', and unknown types use the default.

    def _set_column_widths(self, columns, header_rows=None):
        ws = self.ws
        # Name column gets extra width.
        ws.column_dimensions[get_column_letter(1)].width = 35
        leaf_headers = self._leaf_header_names(columns, header_rows)
        for i, col_def in enumerate(columns[1:], start=2):
            figure_type = col_def.get('figure_type', 'string')
            width = self._width_for(figure_type)
            # Flat legacy exports show ``columns[].name`` directly, so retain
            # their full-period width behaviour. Grouped exports instead show
            # period/analytic labels in merged parent bands: sizing each leaf
            # from the flattened composite name made every numeric column 40
            # characters wide and produced unusable print pagination.
            header = (
                leaf_headers[i - 1]
                if header_rows else str(col_def.get('name') or '')
            )
            width = max(width, min(len(header) + 2, 40))
            ws.column_dimensions[get_column_letter(i)].width = width

    @staticmethod
    def _leaf_header_names(columns, header_rows):
        """Return the deepest visible header label for each physical column."""
        names = [str(column.get('name') or '') for column in columns]
        if not header_rows:
            return names
        for header_row in header_rows:
            for header_cell in header_row:
                start = header_cell['start']
                stop = start + header_cell['colspan']
                for column_index in range(start, stop):
                    names[column_index] = header_cell['name']
        return names

    def _grouped_print_bands(self, columns, header_rows):
        """Return logical value-column bands that fit a readable page.

        The deepest grouped row is the most useful pagination boundary: for
        Trial Balance it is the six-measure analytic block, while two-row
        period/analytic reports use their period groups. Consecutive smaller
        groups share a page when their visible widths fit the same budget.
        """
        if not header_rows or len(columns) <= 1:
            return []
        groups = self._logical_print_groups(columns, header_rows)
        bands = []
        current = None
        current_width = 0.0
        for start, stop in groups:
            group_width = self._value_range_width(start, stop)
            if group_width > _PRINT_VALUE_WIDTH_PER_PAGE:
                if current is not None:
                    bands.append(current)
                    current = None
                    current_width = 0.0
                bands.extend(self._split_value_range(start, stop))
                continue
            if (
                current is not None
                and current_width + group_width > _PRINT_VALUE_WIDTH_PER_PAGE
            ):
                bands.append(current)
                current = None
                current_width = 0.0
            if current is None:
                current = [start, stop]
            else:
                current[1] = stop
            current_width += group_width
        if current is not None:
            bands.append(current)
        return [tuple(band) for band in bands]

    @staticmethod
    def _logical_print_groups(columns, header_rows):
        """Choose the deepest complete row containing grouped value cells."""
        value_stop = len(columns)
        for target_row in range(len(header_rows) - 1, -1, -1):
            groups = []
            for anchor_row, header_row in enumerate(header_rows):
                for header_cell in header_row:
                    if not (
                        anchor_row <= target_row
                        < anchor_row + header_cell['rowspan']
                    ):
                        continue
                    start = max(1, header_cell['start'])
                    stop = min(
                        value_stop,
                        header_cell['start'] + header_cell['colspan'],
                    )
                    if start < stop:
                        groups.append((start, stop))
            groups.sort()
            cursor = 1
            complete = True
            for start, stop in groups:
                if start != cursor:
                    complete = False
                    break
                cursor = stop
            if (
                complete and cursor == value_stop
                and any(stop - start > 1 for start, stop in groups)
            ):
                return groups
        return [
            (column_index, column_index + 1)
            for column_index in range(1, value_stop)
        ]

    def _value_range_width(self, start, stop):
        return sum(
            float(
                self.ws.column_dimensions[
                    get_column_letter(column_index + 1)
                ].width or self._width_for('monetary')
            )
            for column_index in range(start, stop)
        )

    def _split_value_range(self, start, stop):
        bands = []
        band_start = start
        band_width = 0.0
        for column_index in range(start, stop):
            width = self._value_range_width(column_index, column_index + 1)
            if (
                column_index > band_start
                and band_width + width > _PRINT_VALUE_WIDTH_PER_PAGE
            ):
                bands.append((band_start, column_index))
                band_start = column_index
                band_width = 0.0
            band_width += width
        bands.append((band_start, stop))
        return bands

    @staticmethod
    def _split_header_rows_for_print(header_rows, print_bands):
        """Duplicate parent labels where a merged heading crosses a page."""
        if not header_rows or len(print_bands) <= 1:
            return header_rows
        boundaries = {
            stop for _start, stop in print_bands[:-1]
        }
        split_rows = []
        for header_row in header_rows:
            split_row = []
            for header_cell in header_row:
                start = header_cell['start']
                stop = start + header_cell['colspan']
                cuts = [
                    start,
                    *sorted(
                        boundary for boundary in boundaries
                        if start < boundary < stop
                    ),
                    stop,
                ]
                for segment_start, segment_stop in zip(cuts, cuts[1:]):
                    split_cell = dict(header_cell)
                    split_cell['start'] = segment_start
                    split_cell['colspan'] = segment_stop - segment_start
                    split_row.append(split_cell)
            split_rows.append(split_row)
        return split_rows

    def _configure_print_layout(
            self, columns, header_first_row, header_last_row,
            header_rows=None, print_bands=()):
        """Apply deterministic, legible print settings to report sheets."""
        if not columns:
            return
        ws = self.ws
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.25
        ws.page_margins.bottom = 0.25
        ws.page_margins.header = 0.1
        ws.page_margins.footer = 0.1
        ws.print_title_rows = '$%s:$%s' % (
            header_first_row, header_last_row,
        )
        ws.print_title_cols = '$A:$A'

        if header_rows:
            # Explicit logical breaks survive both Excel and LibreOffice.
            # Fit-to-N-pages does not: Calc can ignore the repeated label
            # column and cut straight through merged parent headings.
            ws.sheet_properties.pageSetUpPr.fitToPage = False
            ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
            ws.page_setup.fitToWidth = None
            ws.page_setup.fitToHeight = None
            ws.page_setup.scale = 90
            ws.col_breaks.brk = ()
            for _start, stop in print_bands[:-1]:
                ws.col_breaks.append(Break(id=stop))
            return

        # Preserve the legacy flat-header layout and its full-period column
        # widths. There are no merged group labels to align, so bounded
        # fit-to-width pagination remains appropriate.
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = 0
        value_width = self._value_range_width(1, len(columns))
        ws.page_setup.fitToWidth = max(
            1, int(math.ceil(value_width / _PRINT_VALUE_WIDTH_PER_PAGE)),
        )

    @staticmethod
    def _width_for(figure_type):
        if figure_type in ('monetary', 'float'):
            return 16
        if figure_type == 'integer':
            return 12
        if figure_type == 'percentage':
            return 12
        if figure_type in ('date', 'datetime'):
            return 16
        return 22

    # ---- serialisation ----

    def _serialise(self):
        buf = io.BytesIO()
        self.workbook.save(buf)
        return buf.getvalue()
